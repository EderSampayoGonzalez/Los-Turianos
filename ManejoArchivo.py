from openpyxl import Workbook
from openpyxl import load_workbook
import json

def CrearExcel(DatosPok):
    if DatosPok != list():
        #print(DatosPok)
        libro = Workbook()
        hoja = libro.active
        hoja.title = "pokemon_stats"
        fila = 1
        columna = 0

        for nomb in DatosPok[0]['stats'].keys():
            hoja.cell(fila+1,1,nomb)
            fila+=1

        for pok in DatosPok:
            columna+=1
            fila = 2
            hoja.cell(1,columna+1,pok['name'])
            for y in pok['stats'].values():
                hoja.cell(fila,columna+1,y)
                fila+=1
        try:
            libro.save("Pokemon.xlsx")
        except Exception:
            print("Ocurrió un error, no se pudo guardar el Excel")
    else:
        print("sin")
    
def CrearArchivo(ruta, DatosPoke):
    with open(ruta,"w") as archivo:
        for pokemon in DatosPoke:
            dicc = str(pokemon) + "\n"
            archivo.write(dicc.replace("'",'"'))
    print("Pokemon guardado")
    CrearExcel(DatosPoke)

def LeerArchivo(ruta):
    datospoke = list()
    dicpok = dict()
    try:
        with open(ruta, 'r') as archivo:
            datos = archivo.readlines()
    except NameError:
        print("No se pudo importar el equipo")
    else:
        if datos == list():
            print("No hay datos")
        else:
            print("Equipo importado")
            i = 1
            for diccionario in datos:
                name = diccionario.replace("\n","")
                poke = json.loads(name)
                print(" ", i, poke['name'])
                datospoke.append(poke)
                i+=1
    finally:
        return datospoke

def LeerExcel(ruta):
    DatosStat = list()
    
    datoex = load_workbook(ruta)
    datos = datoex.active

    #print(datos[2][0].value)

    for columna in datos.iter_cols(2,datos.max_column):
        #print(columna[0].value)
        statspok = dict()
        pok = dict()
        pok['name'] = columna[0].value
        for fila in range(1,datos.max_row):
            #print(columna[fila].value)
            nombstat = str(datos[fila+1][0].value)
            statspok[nombstat] = columna[fila].value
        pok['stats'] = statspok
        DatosStat.append(pok)
        
    return DatosStat
        
            

